import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from datetime import datetime, timedelta
from scipy.stats import linregress
from tqdm import tqdm
import warnings
import os

warnings.filterwarnings('ignore')

# ==========================================================
# AYARLAR
# ==========================================================
STOX_FILE = 'top_endeks_hisseleri.txt'
DATA_CACHE_FILE = 'bist_data_cache.pkl'
START_CAPITAL = 100000
COMMISSION_RATE = 0.002
REBALANCE_FREQ = '7D'  # 7 Günlük Periyot
LOOKBACK_DAYS = 20  # Regresyon için geriye dönük gün sayısı
MIN_R_SQUARED = 0.60 # Regresyon uyum kalitesi (0-1 arası)
MIN_SLOPE = 0.0300  # Günlük asgari büyüme hızı
STOP_LOSS_RATE = 0.02 # %10 Stop Loss
MAX_ATR_PERCENT = 0.6    # Yüzdesel oynaklık limiti (ATR/Fiyat)
SLOPE_STOP_FACTOR = 0.000 # Günlük Asgari Getiri Oranı (Örn: 0.005 = Günlük %0.5 artış beklentisi)


def get_tickers_from_file(file_path):
    if not os.path.exists(file_path):
        print(f"HATA: {file_path} dosyası bulunamadı!")
        return []
    with open(file_path, 'r', encoding='utf-8') as f:
        return [t.strip().upper() + ".IS" for t in f.read().splitlines() if t.strip()]

def load_data(tickers, force_refresh=False):
    sim_start = (datetime.now() - timedelta(days=365)).replace(day=1)
    download_start = (sim_start - timedelta(days=400)).strftime('%Y-%m-%d')
    
    cache_exists = os.path.exists(DATA_CACHE_FILE)
    if cache_exists and not force_refresh:
        # Cache dosyasının son değiştirilme tarihini kontrol et
        mtime = datetime.fromtimestamp(os.path.getmtime(DATA_CACHE_FILE))
        # Eğer bugün güncellendiyse cache'den yükle
        if mtime.date() == datetime.now().date():
            data = pd.read_pickle(DATA_CACHE_FILE)
            if hasattr(data, 'columns') and 'Volume' in data.columns:
                print("Güncel veriler cache'den yüklendi.")
                return data
    
    print(f"Yeni veriler indiriliyor (yfinance)... {'(Force Refresh)' if force_refresh else ''}")
    data = yf.download(tickers, start=download_start, auto_adjust=True)
    
    if not data.empty:
        # Hatalı/Eksik hisseleri tespit et
        if isinstance(data.columns, pd.MultiIndex):
            closes = data['Close']
        else:
            closes = data
            
        # Tüm sütunları NaN olan veya hiç inmeyen hisseleri bul
        downloaded_tickers = closes.columns.tolist()
        failed_tickers = [t for t in tickers if t not in downloaded_tickers or closes[t].isna().all()]
        
        if failed_tickers:
            print(f"\n{len(failed_tickers)} hisse indirilemedi, tek tek deneniyor...")
            for t in tqdm(failed_tickers, desc="Eksik veriler tamamlanıyor"):
                try:
                    # Tekli indirme (threads=False daha stabil)
                    t_data = yf.download(t, start=download_start, auto_adjust=True, progress=False, threads=False)
                    if not t_data.empty:
                        if isinstance(data.columns, pd.MultiIndex):
                            # MultiIndex yapısına uygun şekilde ekle
                            for col in ['Close', 'Open', 'High', 'Low', 'Volume']:
                                if col in t_data.columns:
                                    data.loc[:, (col, t)] = t_data[col]
                        else:
                            data[t] = t_data['Close']
                except Exception as e:
                    # print(f"{t} hata: {e}")
                    pass
        
        # Sonuçları kaydet
        data.sort_index(axis=1, inplace=True)
        data.to_pickle(DATA_CACHE_FILE)
    
    return data

def get_clean_data(data):
    """
    Verideki terminal NaN satırlarını temizler.
    Eğer son satırda çok fazla NaN varsa, o satırı siler.
    """
    if data is None or data.empty:
        return data
    
    if isinstance(data.columns, pd.MultiIndex):
        closes = data['Close']
    else:
        closes = data
        
    # Son satırdaki NaN oranını kontrol et
    while not closes.empty:
        last_row_nans = closes.iloc[-1].isna().sum()
        total_cols = len(closes.columns)
        
        # Eğer %50'den fazlası NaN ise bu satırı geçersiz say (örün market henüz açılmamış veya veri eksik)
        if last_row_nans > total_cols * 0.5:
            data = data.iloc[:-1]
            if isinstance(data.columns, pd.MultiIndex):
                closes = data['Close']
            else:
                closes = data
        else:
            break
            
    return data

def get_vectorized_metrics(all_data, lookback_days):
    """
    Tüm hisseler için regresyon ve ATR metriklerini vektörize olarak hesaplar.
    O(n) karmaşıklığında rolling window kullanarak performansı artırır.
    """
    if isinstance(all_data.columns, pd.MultiIndex):
        closes = all_data['Close'].dropna(axis=1, how='all')
        highs = all_data['High']
        lows = all_data['Low']
        volumes = all_data['Volume']
    else:
        closes = all_data
        highs = all_data.get('High', closes)
        lows = all_data.get('Low', closes)
        volumes = all_data.get('Volume', pd.DataFrame(0, index=closes.index, columns=closes.columns))

    n = lookback_days
    log_closes = np.log(closes)
    
    # X değerleri (0,1,2...,n-1)
    x = np.arange(n)
    sum_x = x.sum()
    sum_x2 = (x**2).sum()
    stdev_x = x.std()
    
    # Rolling Sums
    sum_y = log_closes.rolling(window=n).sum()
    sum_xy = log_closes.rolling(window=n).apply(lambda y: (x * y).sum(), raw=True)
    
    # 1. Slope (Eğim) Hesaplama
    # slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x**2)
    slopes = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x**2)
    
    # 2. Intercept (Kesişim) Hesaplama
    # intercept = (sum_y - slope * sum_x) / n
    intercepts = (sum_y - slopes * sum_x) / n
    
    # 3. R-Squared (Uyum Kalitesi) Hesaplama
    # Korelasyon katsayısı r = (n*sum_xy - sum_x*sum_y) / sqrt((n*sum_x2 - sum_x^2) * (n*sum_y2 - sum_y^2))
    sum_y2 = (log_closes**2).rolling(window=n).sum()
    numerator = (n * sum_xy - sum_x * sum_y)
    denominator = np.sqrt((n * sum_x2 - sum_x**2) * (n * sum_y2 - sum_y**2))
    r_squareds = (numerator / denominator)**2
    
    # 4. ATR Oranı (%) Hesaplama
    pc = closes.shift(1)
    tr = np.maximum(
        highs - lows,
        np.maximum((highs - pc).abs(), (lows - pc).abs())
    )
    
    atr = tr.rolling(14).mean()
    atr_pcts = atr / closes
    
    # 5. İskonto (%) Hesaplama
    # expected_log_price = intercept + slope * (n - 1)
    expected_log_prices = intercepts + slopes * (n - 1)
    expected_prices = np.exp(expected_log_prices)
    discounts = (expected_prices - closes) / expected_prices
    
    vol_avgs = volumes.rolling(21).mean()
    
    # Forward fill (ffill) ekleyerek terminal NaN'ları (örneğin seans başı/sonu veri eksikliği) tolere ediyoruz.
    # Bu, Tab 3'teki .dropna() yöntemine benzer şekilde her hissenin en son valid verisini kullanmasını sağlar.
    limit = 5
    return {
        'slopes': slopes.ffill(limit=limit),
        'intercepts': intercepts.ffill(limit=limit),
        'r2': r_squareds.ffill(limit=limit),
        'atr_pct': atr_pcts.ffill(limit=limit),
        'discounts': discounts.ffill(limit=limit),
        'volumes': volumes.ffill(limit=limit),
        'vol_avgs': vol_avgs.ffill(limit=limit),
        'prices': closes.ffill(limit=limit)
    }

from arch import arch_model

def calculate_garch_sigma(prices):
    """
    GARCH(1,1) volatisini hesaplar.
    """
    try:
        returns = 100 * prices.pct_change().dropna() # Yüzdesel getiri
        if len(returns) < 30: # Yetersiz veri
            return np.nan
            
        model = arch_model(returns, vol='Garch', dist='skewt', p=1, q=1)
        res = model.fit(disp='off', show_warning=False)
        
        # Son günün koşullu volatilitesi (yüzde cinsinden döner, tekrar ondalığa çeviriyoruz)
        sigma = res.conditional_volatility.iloc[-1] / 100
        return sigma
    except Exception as e:
        return np.nan

def find_best_candidate(target_date, all_data, lookback_days=LOOKBACK_DAYS, max_atr_percent=MAX_ATR_PERCENT, 
                        min_slope=MIN_SLOPE, min_r2=MIN_R_SQUARED, precalc=None):
    # Eğer precalc varsa, vektörize veriden direkt çek
    if precalc is not None:
        try:
            # En yakın tarihi bul (Borsa tatil olabilir)
            idx = precalc['prices'].index.get_indexer([target_date], method='pad')[0]
            if idx < 0: return []
            
            dt = precalc['prices'].index[idx]
            
            # Tüm hisseler için maske oluştur
            mask = (precalc['slopes'].iloc[idx] > min_slope) & \
                   (precalc['r2'].iloc[idx] > min_r2) & \
                   (precalc['atr_pct'].iloc[idx] <= max_atr_percent)
            
            valid_tickers = mask[mask].index.tolist()
            candidates = []
            
            for ticker in valid_tickers:
                candidates.append({
                    't': ticker,
                    'slope': precalc['slopes'].at[dt, ticker],
                    'r2': precalc['r2'].at[dt, ticker],
                    'score': precalc['discounts'].at[dt, ticker],
                    'price': precalc['prices'].at[dt, ticker],
                    'vol_avg': precalc['vol_avgs'].at[dt, ticker],
                    'vol_curr': precalc['volumes'].at[dt, ticker]
                })
            
            # İlk etapta Skora göre sırala ve ilk 30 adayı al (GARCH hesaplama yükünü azaltmak için)
            candidates.sort(key=lambda x: x['score'], reverse=True)
            top_candidates = candidates[:30]
            
            final_candidates = []
            
            # Top adaylar için GARCH hesapla
            # Veriye erişim için all_data'yı kullanmamız lazım
            if isinstance(all_data.columns, pd.MultiIndex):
                closes = all_data['Close']
            else:
                closes = all_data
            
            for cand in top_candidates:
                ticker = cand['t']
                # Son 1 yılın verisi (veya yeterince uzun)
                # target_date öncesi veriyi al
                try:
                    price_series = closes[ticker].loc[:dt].tail(252) # Son 1 yıl
                    sigma = calculate_garch_sigma(price_series)
                    
                    if not np.isnan(sigma) and sigma > 0:
                        cand['garch_sigma'] = sigma
                        # GARCH Adjusted Score: Score / Sigma 
                        # Yüksek Skor (İskonto) ve Düşük Oynaklık istiyoruz.
                        cand['garch_score'] = cand['score'] / sigma
                        final_candidates.append(cand)
                except KeyError:
                    continue
            
            # GARCH Skora göre sırala (En yüksek)
            final_candidates.sort(key=lambda x: x['garch_score'], reverse=True)
            return final_candidates
            
        except Exception as e:
            # print(f"Vectorized search error: {e}")
            return []

    # Eğer precalc None ise veya fonksiyon buraya düşerse boş liste dön
    return []

if __name__ == "__main__":
    # 1. VERİ YÜKLEME
    tickers = get_tickers_from_file(STOX_FILE)
    if not tickers:
        exit()
    
    all_data = load_data(tickers)
    
    # Veri Ayrıştırma (Simülasyon İçin Loop Dışında Lazım)
def run_simulation(all_data, lookback_days=LOOKBACK_DAYS, min_slope=MIN_SLOPE, min_r2=MIN_R_SQUARED, 
                   stop_loss_rate=STOP_LOSS_RATE, slope_stop_factor=SLOPE_STOP_FACTOR,
                   rebalance_freq=REBALANCE_FREQ, start_capital=START_CAPITAL, commission_rate=COMMISSION_RATE,
                   max_atr_percent=MAX_ATR_PERCENT, progress_callback=None, silent=False, metrics_collector=None):
                   
    # Veri Ayrıştırma
    if isinstance(all_data.columns, pd.MultiIndex):
        try:
            raw_data = all_data['Close'].dropna(axis=1, how='all')
            # raw_volume = all_data['Volume']
        except KeyError:
            raw_data = all_data.xs('Close', axis=1, level=0).dropna(axis=1, how='all')
            # raw_volume = all_data.xs('Volume', axis=1, level=0)
    else:
        raw_data = all_data
            
    sim_start_date = (datetime.now() - timedelta(days=365)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Simülasyon aralığındaki tüm *işlem günlerini* al
    if not isinstance(raw_data.index, pd.DatetimeIndex):
        raw_data.index = pd.to_datetime(raw_data.index)
        
    trading_days = raw_data.loc[sim_start_date:].index
    
    # Setup Variables
    daily_vals = pd.Series(index=trading_days, dtype=float)
    trade_history = []
    current_cash = start_capital
    active_portfolio = [] # [{'t': ticker, 'l': lots, 'b': buy_price, ...}]
    
    # Max Holding Time Parse
    try:
        max_hold_delta = pd.to_timedelta(rebalance_freq)
    except:
        max_hold_delta = pd.Timedelta(days=7)

    # Vektörize Metrikleri Hesapla
    precalc = get_vectorized_metrics(all_data, lookback_days)
    
    total_days = len(trading_days)

    for i, dt in tqdm(enumerate(trading_days), total=total_days, desc="Daily Simulation", disable=silent):
        if progress_callback:
            progress_callback((i + 1) / total_days)
            
        # 1. MEVCUT PORTFÖY KONTROLÜ (STOP & TIME STOP)
        sold_today = False
        
        for item in active_portfolio[:]:
            current_price = raw_data.at[dt, item['t']]
            
            if pd.isna(current_price) or current_price <= 0:
                continue 
            
            should_sell = False
            reason = ""
            
            # A. Zaman Stopu (Max Holding Period)
            time_held = dt - item['buy_dt']
            
            if time_held >= max_hold_delta:
                should_sell = True
                reason = f"SÜRE DOLDU ({rebalance_freq})"
            
            # B. Stop Loss & Trailing Stop
            if not should_sell:
                if current_price > item['max_p']:
                    item['max_p'] = current_price
                    
                if current_price <= item['max_p'] * (1 - stop_loss_rate):
                    should_sell = True
                    reason = "STOP LOSS"
            
            # C. Slope Stop (Minimum Getiri Stopu)
            if not should_sell and slope_stop_factor != 0:
                floor_price = item['b'] * ((1 + slope_stop_factor) ** item['trading_days_held'])
                if current_price < floor_price:
                    should_sell = True
                    reason = "SLOPE STOP"
                    
            # D. Hacim Stopu (Volume Stop)
            if not should_sell:
                try:
                    curr_vol = precalc['volumes'].at[dt, item['t']]
                    avg_vol = precalc['vol_avgs'].at[dt, item['t']]
                    if not pd.isna(curr_vol) and not pd.isna(avg_vol) and avg_vol > 0:
                        if curr_vol*(4.5) < avg_vol:
                            should_sell = True
                            reason = "VOLUME STOP"
                except:
                    pass

            # D. SATIŞ İŞLEMİ
            if should_sell:
                revenue = item['l'] * current_price * (1 - commission_rate)
                current_cash += revenue
                
                pl_pct = (current_price / item['b'] - 1) * 100
                trade_history.append([
                    dt.strftime('%Y-%m-%d'), 
                    item['t'].replace('.IS',''), 
                    item['l'],
                    f"{current_price:.2f}", 
                    reason, 
                    f"{current_cash:,.2f}", 
                    f"P/L: %{pl_pct:.2f} | Peak: {item['max_p']:.2f}"
                ])
                
                if metrics_collector is not None:
                    metrics_collector.append({
                        'pl_pct': pl_pct,
                        'slope': item['slope'],
                        'r2': item.get('r2', 0),
                        'score': item.get('score', 0)
                    })
                
                active_portfolio.remove(item)
                sold_today = True
            else:
                item['trading_days_held'] += 1

        # 2. ALIM İŞLEMİ (Eğer portföy boşsa ve bugün satış yapılmadıysa)
        if not active_portfolio and not sold_today:
            candidates = find_best_candidate(dt, all_data, lookback_days, max_atr_percent, min_slope, min_r2, precalc=precalc)
            
            if candidates:
                top_pick = candidates[0]
                buy_price = top_pick['price']
                
                if buy_price > 0:
                    lots = int(current_cash / buy_price)
                    if lots > 0:
                        cost = lots * buy_price
                        current_cash -= cost * (1 + commission_rate)
                        
                        active_portfolio.append({
                            't': top_pick['t'],
                            'l': lots,
                            'b': buy_price,
                            'max_p': buy_price,
                            'slope': top_pick['slope'],
                            'r2': top_pick['r2'],
                            'score': top_pick['score'],
                            'buy_dt': dt,                  
                            'trading_days_held': 0         
                        })
                        
                        trade_history.append([
                            dt.strftime('%Y-%m-%d'), 
                            top_pick['t'].replace('.IS',''), 
                            lots,
                            f"{buy_price:.2f}", 
                            "ALIS", 
                            f"{current_cash:,.2f}", 
                            f"(Eğim: {top_pick['slope']:.4f}, R2: {top_pick['r2']:.2f})"
                        ])

        # 3. GÜNLÜK DEĞERLEME KAYDI
        port_val = current_cash
        for item in active_portfolio:
            cp = raw_data.at[dt, item['t']]
            if pd.isna(cp) or cp <= 0:
                cp = item['b']
            port_val += item['l'] * cp
            
        daily_vals[dt] = port_val

    # Sonucu Döndür
    return daily_vals, trade_history, daily_vals.iloc[-1]

if __name__ == "__main__":
    # 1. VERİ YÜKLEME
    tickers = get_tickers_from_file(STOX_FILE)
    if not tickers:
        exit()
    
    all_data = load_data(tickers)
    
    # 2. SİMÜLASYON BAŞLAT
    print(f"--- Simülasyon Başlıyor (Regresyon: {LOOKBACK_DAYS} Gün) ---")
    daily_vals, trade_history, final_balance = run_simulation(all_data)


    # SON DURUM RAPORU
    final_balance = daily_vals.iloc[-1]
    roi = ((final_balance - START_CAPITAL) / START_CAPITAL) * 100
    
    print(f"\n🎯 Sonuç: {START_CAPITAL:,.0f} TL -> {final_balance:,.2f} TL")
    print(f"Toplam Getiri: %{roi:.2f}")
    
    # EXCEL ÇIKTISI (Renklendirilmiş)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        if trade_history:
            columns = ["Tarih", "Hisse", "Lot", "Fiyat", "İşlem", "Nakit", "Bilgi"]
            df_history = pd.DataFrame(trade_history, columns=columns)
            
            # Excel dosyası oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "İşlem Geçmişi"
            
            # Başlıkları ekle
            ws.append(columns)
            
            # Verileri ekle ve renklendir
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Açık Yeşil
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Açık Kırmızı
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Beyaz
            
            for index, row in df_history.iterrows():
                ws.append(row.tolist())
                current_row = ws[index + 2] # +2 çünkü başlık var ve openpyxl 1-indexed
                
                # Renklendirme Mantığı
                is_buy = row['İşlem'] == 'ALIS'
                info = row['Bilgi']
                action = str(row['İşlem'])
                
                fill = None
                if is_buy:
                    fill = white_fill
                elif 'P/L: %-' in info: # Zarar
                     fill = red_fill
                elif 'P/L: %' in info: # Kar
                     fill = green_fill
                elif 'SÜRE DOLDU ' in action:
                     fill = green_fill
                
                if fill:
                    for cell in current_row:
                        cell.fill = fill
                        
            wb.save('trade_history.xlsx')
            print(f"\nİşlem geçmişi 'trade_history.xlsx' dosyasına renkli olarak kaydedildi.")
            
    except ImportError:
        # Fallback to CSV if openpyxl is missing
        if trade_history:
            columns = ["Tarih", "Hisse", "Lot", "Fiyat", "İşlem", "Nakit", "Bilgi"]
            df_history = pd.DataFrame(trade_history, columns=columns)
            df_history.to_csv('trade_history.csv', index=False, encoding='utf-8-sig', sep=';')
            print(f"\nİşlem geçmişi 'trade_history.csv' dosyasına kaydedildi (Excel renkli çıktı için openpyxl kurun).")


    # --- GÖRSELLEŞTİRME ---
    plt.style.use('dark_background')
    fig = plt.figure(figsize=(14, 10))
    gs = fig.add_gridspec(2, 1, height_ratios=[2, 1])
    
    # Ana Grafik
    ax1 = fig.add_subplot(gs[0])
    ax1.plot(daily_vals.index, daily_vals.values, color='cyan', lw=2, label="Exponential Regression Model")
    ax1.axhline(y=START_CAPITAL, color='white', ls='--', alpha=0.3)
    ax1.set_title(f"Üslü Regresyon Stratejisi | Lookback: {LOOKBACK_DAYS} Gün | R2 >= {MIN_R_SQUARED}", fontsize=14)
    ax1.yaxis.set_major_formatter(mtick.StrMethodFormatter('{x:,.0f} TL'))
    ax1.legend()
    ax1.grid(True, alpha=0.15)
    
    # Scrollable Table Implementation
    ax2 = fig.add_subplot(gs[1])
    ax2.axis('off')
    
    # Veri Hazırlığı
    columns = ["Tarih", "Hisse", "Lot", "Fiyat", "İşlem", "Nakit", "Bilgi"]
    if not trade_history:
        trade_history = [["-"] * 7]
    
    # Tabloyu tersten sırala (En yeni en üstte)
    table_data_full = list(reversed(trade_history))
    
    rows_per_page = 12
    total_rows = len(table_data_full)
    
    # İlk sayfa verisi
    current_data = table_data_full[:rows_per_page]
    # Eğer veri azsa boş satırlarla doldur
    while len(current_data) < rows_per_page:
        current_data.append([""] * len(columns))
        
    the_table = ax2.table(cellText=current_data, colLabels=columns, loc='center', cellLoc='center', bbox=[0, 0, 1, 1])
    the_table.auto_set_font_size(False)
    the_table.set_fontsize(9)
    
    # Renklendirme Fonksiyonu
    def color_rows(data_slice):
        for i, row in enumerate(data_slice):
            # Hücrelerin varsayılan rengi (Dark Theme)
            bg_color = 'none' 
            text_color = 'white'
            
            if row[0] != "" and row[0] != "-":
                # İşlem tipine göre renk
                action = str(row[4]) # İşlem Sütunu (str cast önemli)
                info = row[6]   # Bilgi Sütunu
                
                if action == 'ALIS':
                    bg_color = '#ffffff' # Beyaz Arkaplan
                    text_color = 'black' # Siyah Yazı
                elif 'P/L: %-' in info: # Zarar
                    bg_color = '#c0392b' # Kırmızı
                    text_color = 'white'
                elif 'P/L: %' in info: # Kar (Pozitif)
                    bg_color = '#27ae60' # Yeşil
                    text_color = 'white'
                elif 'SÜRE DOLDU' in action or 'VOLUME STOP' in action:
                    bg_color = '#27ae60' # Yeşil
                    text_color = 'white'
            
            # Satırdaki tüm hücreleri boya
            for j in range(len(columns)):
                cell = the_table[i+1, j] # +1 header olduğu için
                cell.set_facecolor(bg_color)
                cell.set_text_props(color=text_color)
    
    color_rows(current_data)
    
    # Slider Ekleme
    if total_rows > rows_per_page:
        from matplotlib.widgets import Slider
        
        # Slider ekseni
        ax_slider = plt.axes([0.92, 0.15, 0.02, 0.25], facecolor='lightgoldenrodyellow')
        slider = Slider(
            ax=ax_slider,
            label="Scroll",
            valmin=0,
            valmax=total_rows - rows_per_page,
            valinit=0,
            orientation="vertical",
            valstep=1
        )
        
        def update(val):
            start = int(slider.val) # Slider yukarıdan aşağıya çalışsın diye ters çevirmedik, data zaten ters
            end = start + rows_per_page
            new_data = table_data_full[start:end]
            
            # Veri azsa doldur
            while len(new_data) < rows_per_page:
                new_data.append([""] * len(columns))
            
            # Hücre içeriklerini güncelle
            for i, row in enumerate(new_data):
                for j, val in enumerate(row):
                    the_table[i+1, j].get_text().set_text(val)
            
            # Renkleri güncelle
            color_rows(new_data)
            fig.canvas.draw_idle()
            
        slider.on_changed(update)

    plt.tight_layout(rect=[0, 0, 0.9, 1]) # Slider için yer aç
    plt.show()
